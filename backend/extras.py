"""
Extras module: repair photos, cash-register movements, and Excel report exports.

All routes are mounted at /api/*.  Owner-gated where destructive.
"""
from __future__ import annotations

import io
import uuid
import logging
from datetime import datetime, timezone
from typing import Optional, List

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from fastapi.responses import Response, StreamingResponse
from pydantic import BaseModel, Field
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

import backup as _backup  # reuse init_storage/put/get from backup module

logger = logging.getLogger(__name__)

APP_PREFIX = "pitstock/repair-photos"
MAX_PHOTO_BYTES = 5 * 1024 * 1024
MAX_PHOTOS_PER_CARD = 10
ALLOWED_TYPES = {"image/jpeg", "image/png", "image/webp", "image/heic", "image/heif"}


# ---------- Models ----------

class CashMovement(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    date: str  # YYYY-MM-DD (business day)
    direction: str  # IN | OUT
    amount: float = Field(gt=0)
    payment_method_id: Optional[str] = None
    payment_method_name: str = ""
    category: str = "other"  # deposit | withdrawal | expense | other
    note: str = ""
    created_by: str = ""
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())


class CashMovementCreate(BaseModel):
    date: Optional[str] = None
    direction: str  # IN | OUT
    amount: float = Field(gt=0)
    payment_method_id: Optional[str] = None
    category: str = "other"
    note: str = ""


# ---------- Excel helpers ----------

def _xlsx_response(wb: Workbook, filename: str) -> StreamingResponse:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _style_header(ws, headers: List[str]):
    ws.append(headers)
    fill = PatternFill("solid", fgColor="1E3A8A")
    font = Font(bold=True, color="FFFFFF")
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.fill = fill
        c.font = font
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22


def _autosize(ws):
    for col_cells in ws.columns:
        col_letter = col_cells[0].column_letter
        max_len = max((len(str(c.value)) if c.value is not None else 0 for c in col_cells), default=10)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 45)


# ---------- Router builder ----------

def register(db, current_user_dep, require_owner_dep) -> APIRouter:
    router = APIRouter()

    # ============ Repair photos ============

    @router.post("/repairs/{rid}/photos")
    async def upload_repair_photo(
        rid: str,
        file: UploadFile = File(...),
        kind: str = Form("general"),
        caption: str = Form(""),
        user: dict = Depends(current_user_dep),
    ):
        card = await db.repairs.find_one({"id": rid}, {"_id": 0})
        if not card:
            raise HTTPException(status_code=404, detail="Repair card not found")
        if len(card.get("photos") or []) >= MAX_PHOTOS_PER_CARD:
            raise HTTPException(status_code=400, detail=f"This card already has {MAX_PHOTOS_PER_CARD} photos")
        content_type = (file.content_type or "").lower()
        if content_type not in ALLOWED_TYPES:
            raise HTTPException(status_code=400, detail=f"Unsupported image type: {content_type}")
        data = await file.read()
        if len(data) > MAX_PHOTO_BYTES:
            raise HTTPException(status_code=413, detail=f"Photo too large (max {MAX_PHOTO_BYTES // 1024 // 1024} MB)")
        ext = (file.filename.rsplit(".", 1)[-1] if "." in (file.filename or "") else "jpg").lower()
        pid = str(uuid.uuid4())
        path = f"{APP_PREFIX}/{rid}/{pid}.{ext}"
        try:
            result = _backup._put_object(path, data, content_type)
        except Exception as e:
            logger.exception("Photo upload failed")
            raise HTTPException(status_code=502, detail=f"Storage error: {str(e)[:200]}")
        photo = {
            "id": pid,
            "storage_path": result["path"],
            "filename": file.filename or f"{pid}.{ext}",
            "content_type": content_type,
            "size": len(data),
            "caption": caption or "",
            "kind": kind if kind in ("before", "after", "damage", "general") else "general",
            "uploaded_by": user.get("email", ""),
            "uploaded_at": datetime.now(timezone.utc).isoformat(),
        }
        await db.repairs.update_one(
            {"id": rid},
            {"$push": {"photos": photo}, "$set": {"updated_at": photo["uploaded_at"]}},
        )
        return photo

    @router.get("/repairs/{rid}/photos")
    async def list_repair_photos(rid: str, user: dict = Depends(current_user_dep)):
        card = await db.repairs.find_one({"id": rid}, {"_id": 0, "photos": 1})
        if not card:
            raise HTTPException(status_code=404, detail="Repair card not found")
        return card.get("photos") or []

    @router.delete("/repairs/{rid}/photos/{photo_id}")
    async def delete_repair_photo(rid: str, photo_id: str, user: dict = Depends(current_user_dep)):
        card = await db.repairs.find_one({"id": rid}, {"_id": 0, "photos": 1})
        if not card:
            raise HTTPException(status_code=404, detail="Repair card not found")
        photos = card.get("photos") or []
        if not any(p["id"] == photo_id for p in photos):
            raise HTTPException(status_code=404, detail="Photo not found")
        remaining = [p for p in photos if p["id"] != photo_id]
        await db.repairs.update_one({"id": rid}, {"$set": {"photos": remaining, "updated_at": datetime.now(timezone.utc).isoformat()}})
        return {"ok": True}

    @router.get("/photos/{photo_id}")
    async def download_repair_photo(photo_id: str, user: dict = Depends(current_user_dep)):
        # find photo record in any card
        card = await db.repairs.find_one({"photos.id": photo_id}, {"_id": 0, "photos": 1})
        if not card:
            raise HTTPException(status_code=404, detail="Photo not found")
        photo = next((p for p in card["photos"] if p["id"] == photo_id), None)
        if not photo:
            raise HTTPException(status_code=404, detail="Photo not found")
        try:
            data = _backup._get_object(photo["storage_path"])
        except Exception as e:
            raise HTTPException(status_code=502, detail=f"Storage error: {str(e)[:200]}")
        return Response(content=data, media_type=photo.get("content_type", "image/jpeg"), headers={"Cache-Control": "private, max-age=3600"})

    # ============ Cash movements ============

    @router.get("/cash-movements")
    async def list_cash_movements(date: Optional[str] = None, user: dict = Depends(current_user_dep)):
        q = {"date": date} if date else {}
        return await db.cash_movements.find(q, {"_id": 0}).sort("created_at", -1).to_list(2000)

    @router.post("/cash-movements", response_model=CashMovement)
    async def create_cash_movement(payload: CashMovementCreate, user: dict = Depends(current_user_dep)):
        if payload.direction not in ("IN", "OUT"):
            raise HTTPException(status_code=400, detail="direction must be IN or OUT")
        d = payload.date or datetime.now(timezone.utc).strftime("%Y-%m-%d")
        method_name = ""
        if payload.payment_method_id:
            m = await db.payment_methods.find_one({"id": payload.payment_method_id}, {"_id": 0})
            if not m:
                raise HTTPException(status_code=404, detail="Payment method not found")
            method_name = m["name"]
        mv = CashMovement(
            date=d,
            direction=payload.direction,
            amount=round(payload.amount, 2),
            payment_method_id=payload.payment_method_id,
            payment_method_name=method_name,
            category=payload.category or "other",
            note=payload.note or "",
            created_by=user.get("email", ""),
        )
        await db.cash_movements.insert_one(mv.model_dump())
        # Reflect on the payment method ledger too so account balances stay in sync
        if payload.payment_method_id:
            entry = {
                "id": str(uuid.uuid4()),
                "method_id": payload.payment_method_id,
                "method_name": method_name,
                "direction": payload.direction,
                "amount": mv.amount,
                "counterpart": mv.category,
                "note": mv.note or f"Cash-register {payload.category}",
                "source": "cash_movement",
                "source_id": mv.id,
                "date": mv.date,
                "created_by": user.get("email", ""),
                "created_at": mv.created_at,
            }
            await db.payment_entries.insert_one(entry)
        return mv

    @router.delete("/cash-movements/{mid}")
    async def delete_cash_movement(mid: str, user: dict = Depends(current_user_dep)):
        mv = await db.cash_movements.find_one({"id": mid}, {"_id": 0})
        if not mv:
            raise HTTPException(status_code=404, detail="Not found")
        await db.cash_movements.delete_one({"id": mid})
        await db.payment_entries.delete_many({"source": "cash_movement", "source_id": mid})
        return {"ok": True}

    # ============ Excel exports ============

    @router.get("/reports/inventory/excel")
    async def inventory_excel(user: dict = Depends(current_user_dep)):
        items = await db.inventory.find({}, {"_id": 0}).sort("name", 1).to_list(10000)
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventory"
        _style_header(ws, ["SKU", "Barcode", "Name", "Category", "Location", "Qty", "Reorder", "Cost €", "Selling €", "Stock value €"])
        for i in items:
            ws.append([
                i.get("sku", ""), i.get("barcode", ""), i.get("name", ""), i.get("category", ""), i.get("location", ""),
                i.get("quantity", 0), i.get("reorder_point", 0),
                round(float(i.get("cost_price") or 0), 2), round(float(i.get("selling_price") or 0), 2),
                round(float(i.get("cost_price") or 0) * i.get("quantity", 0), 2),
            ])
        _autosize(ws)
        return _xlsx_response(wb, f"inventory-{datetime.now().strftime('%Y%m%d')}.xlsx")

    @router.get("/reports/invoices/excel")
    async def invoices_excel(start: Optional[str] = None, end: Optional[str] = None, user: dict = Depends(current_user_dep)):
        q = {}
        if start or end:
            q["created_at"] = {}
            if start: q["created_at"]["$gte"] = start
            if end: q["created_at"]["$lte"] = end + "T23:59:59"
        invs = await db.invoices.find(q, {"_id": 0}).sort("created_at", -1).to_list(10000)
        wb = Workbook()
        ws = wb.active
        ws.title = "Invoices"
        _style_header(ws, ["Invoice #", "Date", "Customer", "Lines", "Subtotal €", "Tax €", "Total €", "Status", "Payment method", "Paid at", "Note"])
        for i in invs:
            ws.append([
                i.get("invoice_number", ""),
                (i.get("created_at") or "")[:10],
                i.get("customer_name") or "Walk-in",
                len(i.get("lines") or []),
                round(float(i.get("subtotal") or 0), 2),
                round(float(i.get("tax") or 0), 2),
                round(float(i.get("total") or 0), 2),
                i.get("status", ""),
                i.get("payment_method_name") or "",
                (i.get("paid_at") or "")[:16].replace("T", " "),
                (i.get("note") or "")[:200],
            ])
        _autosize(ws)
        return _xlsx_response(wb, f"invoices-{datetime.now().strftime('%Y%m%d')}.xlsx")

    @router.get("/reports/profit/excel")
    async def profit_excel(start: Optional[str] = None, end: Optional[str] = None, user: dict = Depends(current_user_dep)):
        # reuse the existing profit report logic via HTTP-free import: call the DB directly.
        from datetime import datetime as _dt, timedelta as _td
        now = _dt.now(timezone.utc)
        end_dt = _dt.fromisoformat(end).replace(tzinfo=timezone.utc) if end else now
        start_dt = _dt.fromisoformat(start).replace(tzinfo=timezone.utc) if start else (now - _td(days=30))
        end_iso = end_dt.replace(hour=23, minute=59, second=59).isoformat()
        start_iso = start_dt.replace(hour=0, minute=0, second=0).isoformat()
        txns = await db.transactions.find({"type": "OUT", "created_at": {"$gte": start_iso, "$lte": end_iso}}, {"_id": 0}).to_list(20000)
        items_map = {i["id"]: i for i in await db.inventory.find({}, {"_id": 0}).to_list(5000)}
        by_item = {}
        total_rev = total_cost = 0.0
        for t in txns:
            rev = float(t.get("total") or 0)
            cost_unit = float(t.get("item_cost") or 0)
            if cost_unit == 0:
                it = items_map.get(t["item_id"])
                cost_unit = float(it.get("cost_price") or 0) if it else 0
            cost = cost_unit * (t.get("quantity") or 0)
            profit = rev - cost
            total_rev += rev; total_cost += cost
            k = t["item_id"]
            entry = by_item.setdefault(k, {"item_id": k, "sku": t.get("item_sku", ""), "name": t.get("item_name", ""), "category": (items_map.get(k) or {}).get("category", ""), "qty_sold": 0, "revenue": 0, "cost": 0, "profit": 0})
            entry["qty_sold"] += t.get("quantity") or 0
            entry["revenue"] += rev; entry["cost"] += cost; entry["profit"] += profit
        wb = Workbook()
        # summary
        summary = wb.active
        summary.title = "Summary"
        _style_header(summary, ["Metric", "Value"])
        for k, v in [
            ("From", start_dt.strftime("%Y-%m-%d")),
            ("To", end_dt.strftime("%Y-%m-%d")),
            ("Total revenue €", round(total_rev, 2)),
            ("Cost of goods €", round(total_cost, 2)),
            ("Gross profit €", round(total_rev - total_cost, 2)),
            ("Margin %", round((total_rev - total_cost) / total_rev * 100.0, 2) if total_rev else 0),
        ]:
            summary.append([k, v])
        _autosize(summary)
        # by item
        ws = wb.create_sheet("By part")
        _style_header(ws, ["SKU", "Part", "Category", "Sold", "Revenue €", "Cost €", "Profit €", "Margin %"])
        for row in sorted(by_item.values(), key=lambda x: -x["profit"]):
            m = (row["profit"] / row["revenue"] * 100.0) if row["revenue"] else 0
            ws.append([row["sku"], row["name"], row["category"], row["qty_sold"], round(row["revenue"], 2), round(row["cost"], 2), round(row["profit"], 2), round(m, 2)])
        _autosize(ws)
        return _xlsx_response(wb, f"profit-{start_dt.strftime('%Y%m%d')}-{end_dt.strftime('%Y%m%d')}.xlsx")

    @router.get("/reports/cash-register/excel")
    async def cash_register_excel(date: Optional[str] = None, user: dict = Depends(current_user_dep)):
        d = date or datetime.now(timezone.utc).strftime("%Y-%m-%d")
        start = d + "T00:00:00+00:00"; end = d + "T23:59:59+00:00"
        invs = await db.invoices.find({"paid_at": {"$gte": start, "$lte": end}, "status": "paid"}, {"_id": 0}).to_list(2000)
        moves = await db.cash_movements.find({"date": d}, {"_id": 0}).sort("created_at", 1).to_list(2000)
        wb = Workbook()
        ws = wb.active
        ws.title = f"Till {d}"
        _style_header(ws, ["Time", "Type", "Reference", "Customer / Note", "Payment method", "Amount €"])
        for i in invs:
            ws.append([
                (i.get("paid_at") or "")[11:16],
                "Invoice",
                i.get("invoice_number", ""),
                i.get("customer_name") or "Walk-in",
                i.get("payment_method_name") or "",
                round(float(i.get("total") or 0), 2),
            ])
        for m in moves:
            ws.append([
                (m.get("created_at") or "")[11:16],
                f"Cash {m['direction']}",
                m.get("category", ""),
                m.get("note", ""),
                m.get("payment_method_name", ""),
                round(float(m.get("amount") or 0), 2) * (1 if m["direction"] == "IN" else -1),
            ])
        _autosize(ws)
        # summary sheet
        s = wb.create_sheet("Summary")
        _style_header(s, ["Metric", "Value"])
        revenue = round(sum(float(i.get("total") or 0) for i in invs), 2)
        tax = round(sum(float(i.get("tax") or 0) for i in invs), 2)
        m_in = round(sum(float(x["amount"]) for x in moves if x["direction"] == "IN"), 2)
        m_out = round(sum(float(x["amount"]) for x in moves if x["direction"] == "OUT"), 2)
        for k, v in [
            ("Date", d),
            ("Paid invoices", len(invs)),
            ("Revenue €", revenue),
            ("Tax €", tax),
            ("Manual IN €", m_in),
            ("Manual OUT €", m_out),
            ("Net €", round(revenue + m_in - m_out, 2)),
        ]:
            s.append([k, v])
        _autosize(s)
        return _xlsx_response(wb, f"cash-register-{d}.xlsx")

    return router
