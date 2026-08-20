"""Iteration 10 tests: dashboard v2, repair photos, repair BTW/tax, cash movements,
cash-register manual totals, Excel exports, settings default_tax_rate, auth gating."""
import io
import os
import zipfile
from datetime import datetime, timezone

import pytest
import requests

from conftest import BASE_URL

XLSX_CT = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
TODAY = datetime.now(timezone.utc).strftime("%Y-%m-%d")

# Smallest valid 1x1 JPEG
JPEG_1PX = bytes.fromhex(
    "ffd8ffe000104a46494600010100000100010000ffdb004300ffffffffffffffffffffffffffffff"
    "ffffffffffffffffffffffffffffffffffffffffffffffffffffffffffffffffffffffffffffffff"
    "ffffffffffffffffffffffffffffffffffffffffffffffffffffffffffffffffffffffffffffffff"
    "ffffc00011080001000101011100ffc40014000100000000000000000000000000000009ffda0008"
    "01010000013f10"
)
PNG_1PX = bytes.fromhex(
    "89504e470d0a1a0a0000000d4948445200000001000000010806000000"
    "1f15c4890000000a49444154789c6300010000050001"
    "0d0a2db40000000049454e44ae426082"
)


# ---------- helpers ----------

def _open_repair_id(owner_client):
    r = owner_client.get(f"{BASE_URL}/api/repairs", timeout=60)
    assert r.status_code == 200, r.text[:300]
    cards = r.json()
    if cards:
        return cards[0]["id"]
    c = owner_client.post(f"{BASE_URL}/api/repairs", json={
        "customer_name": "TEST_Photo Customer", "car_make": "TEST", "car_model": "X",
        "car_plate": "TEST-001", "complaint": "TEST"}, timeout=60)
    assert c.status_code == 200, c.text[:300]
    return c.json()["id"]


@pytest.fixture(scope="module")
def method_id(owner_client):
    r = owner_client.get(f"{BASE_URL}/api/payment-methods", timeout=60)
    assert r.status_code == 200, r.text[:300]
    methods = r.json()
    if not methods:
        cr = owner_client.post(f"{BASE_URL}/api/payment-methods",
                               json={"name": "TEST_Cash", "type": "cash"}, timeout=60)
        assert cr.status_code in (200, 201), cr.text[:300]
        return cr.json()["id"]
    return methods[0]["id"]


# ================= Dashboard v2 =================
class TestDashboardV2:
    def test_dashboard_summary_new_fields(self, owner_client):
        r = owner_client.get(f"{BASE_URL}/api/dashboard/summary", timeout=90)
        assert r.status_code == 200, r.text[:400]
        d = r.json()
        for k in ("open_cars", "open_cars_count", "revenue_today", "revenue_week",
                  "revenue_month", "mechanic_hours_today", "mechanic_minutes_today"):
            assert k in d, f"missing {k}"
        assert isinstance(d["open_cars"], list)
        assert d["open_cars_count"] == len(d["open_cars"])
        assert isinstance(d["mechanic_hours_today"], list)
        if d["open_cars"]:
            car = d["open_cars"][0]
            for k in ("card_number", "customer_name", "car_make", "car_model", "car_plate",
                      "mechanic_name", "status", "grand_total", "hours_in_shop", "cover_photo_id"):
                assert k in car, f"open_cars missing {k}"
            assert car["status"] in ("open", "in_progress")
            assert isinstance(car["hours_in_shop"], (int, float))
        for m in d["mechanic_hours_today"]:
            assert set(m.keys()) == {"name", "hours"}

    def test_dashboard_requires_auth(self):
        r = requests.get(f"{BASE_URL}/api/dashboard/summary", timeout=60)
        assert r.status_code == 401, r.status_code


# ================= Repair photos =================
class TestRepairPhotos:
    def test_upload_list_download_delete(self, owner_client):
        rid = _open_repair_id(owner_client)
        r = owner_client.post(
            f"{BASE_URL}/api/repairs/{rid}/photos",
            files={"file": ("TEST_before.jpg", JPEG_1PX, "image/jpeg")},
            data={"kind": "before", "caption": "TEST caption"}, timeout=90)
        assert r.status_code == 200, r.text[:400]
        p = r.json()
        assert p["kind"] == "before"
        assert p["size"] == len(JPEG_1PX)
        assert p["storage_path"] and isinstance(p["id"], str)
        assert p["caption"] == "TEST caption"
        pid = p["id"]

        lr = owner_client.get(f"{BASE_URL}/api/repairs/{rid}/photos", timeout=60)
        assert lr.status_code == 200
        assert any(x["id"] == pid for x in lr.json())

        dr = owner_client.get(f"{BASE_URL}/api/photos/{pid}", timeout=90)
        assert dr.status_code == 200, dr.text[:300]
        assert dr.content == JPEG_1PX
        assert dr.headers.get("content-type", "").startswith("image/")

        # card document exposes photos
        cr = owner_client.get(f"{BASE_URL}/api/repairs/{rid}", timeout=60)
        assert cr.status_code == 200
        assert any(x["id"] == pid for x in (cr.json().get("photos") or []))

        rm = owner_client.delete(f"{BASE_URL}/api/repairs/{rid}/photos/{pid}", timeout=60)
        assert rm.status_code == 200, rm.text[:300]
        lr2 = owner_client.get(f"{BASE_URL}/api/repairs/{rid}/photos", timeout=60)
        assert not any(x["id"] == pid for x in lr2.json())
        assert owner_client.get(f"{BASE_URL}/api/photos/{pid}", timeout=60).status_code == 404

    def test_png_allowed(self, owner_client):
        rid = _open_repair_id(owner_client)
        r = owner_client.post(f"{BASE_URL}/api/repairs/{rid}/photos",
                              files={"file": ("TEST_p.png", PNG_1PX, "image/png")},
                              data={"kind": "damage"}, timeout=90)
        assert r.status_code == 200, r.text[:300]
        owner_client.delete(f"{BASE_URL}/api/repairs/{rid}/photos/{r.json()['id']}", timeout=60)

    def test_pdf_rejected_400(self, owner_client):
        rid = _open_repair_id(owner_client)
        r = owner_client.post(f"{BASE_URL}/api/repairs/{rid}/photos",
                              files={"file": ("TEST_x.pdf", b"%PDF-1.4 test", "application/pdf")},
                              data={"kind": "general"}, timeout=60)
        assert r.status_code == 400, f"{r.status_code} {r.text[:200]}"

    def test_oversized_rejected_413(self, owner_client):
        rid = _open_repair_id(owner_client)
        big = JPEG_1PX + os.urandom(6 * 1024 * 1024)
        r = owner_client.post(f"{BASE_URL}/api/repairs/{rid}/photos",
                              files={"file": ("TEST_big.jpg", big, "image/jpeg")},
                              data={"kind": "general"}, timeout=180)
        assert r.status_code == 413, f"{r.status_code} {r.text[:200]}"

    def test_eleventh_photo_rejected_400(self, owner_client):
        cr = owner_client.post(f"{BASE_URL}/api/repairs", json={
            "customer_name": "TEST_Limit", "car_make": "TEST", "car_plate": "TEST-LIM"}, timeout=60)
        assert cr.status_code == 200, cr.text[:300]
        rid = cr.json()["id"]
        try:
            for i in range(10):
                r = owner_client.post(f"{BASE_URL}/api/repairs/{rid}/photos",
                                      files={"file": (f"TEST_{i}.jpg", JPEG_1PX, "image/jpeg")},
                                      data={"kind": "general"}, timeout=90)
                assert r.status_code == 200, f"photo {i}: {r.status_code} {r.text[:200]}"
            r11 = owner_client.post(f"{BASE_URL}/api/repairs/{rid}/photos",
                                    files={"file": ("TEST_11.jpg", JPEG_1PX, "image/jpeg")},
                                    data={"kind": "general"}, timeout=90)
            assert r11.status_code == 400, f"{r11.status_code} {r11.text[:200]}"
        finally:
            owner_client.delete(f"{BASE_URL}/api/repairs/{rid}", timeout=60)

    def test_photo_routes_require_auth(self, owner_client):
        rid = _open_repair_id(owner_client)
        assert requests.get(f"{BASE_URL}/api/repairs/{rid}/photos", timeout=60).status_code == 401
        assert requests.post(f"{BASE_URL}/api/repairs/{rid}/photos",
                             files={"file": ("a.jpg", JPEG_1PX, "image/jpeg")},
                             timeout=60).status_code == 401

    def test_photo_missing_card_404(self, owner_client):
        r = owner_client.get(f"{BASE_URL}/api/repairs/does-not-exist/photos", timeout=60)
        assert r.status_code == 404

    def test_staff_can_upload_and_view(self, owner_client, staff_client):
        rid = _open_repair_id(owner_client)
        r = staff_client.post(f"{BASE_URL}/api/repairs/{rid}/photos",
                              files={"file": ("TEST_staff.jpg", JPEG_1PX, "image/jpeg")},
                              data={"kind": "after"}, timeout=90)
        assert r.status_code == 200, r.text[:300]
        pid = r.json()["id"]
        assert staff_client.get(f"{BASE_URL}/api/photos/{pid}", timeout=60).status_code == 200
        owner_client.delete(f"{BASE_URL}/api/repairs/{rid}/photos/{pid}", timeout=60)


# ================= Repair BTW / tax =================
class TestRepairTax:
    @pytest.fixture(scope="class")
    def card_id(self, owner_client):
        cr = owner_client.post(f"{BASE_URL}/api/repairs", json={
            "customer_name": "TEST_Tax", "car_make": "TEST", "car_plate": "TEST-TAX"}, timeout=60)
        assert cr.status_code == 200, cr.text[:300]
        rid = cr.json()["id"]
        yield rid
        owner_client.delete(f"{BASE_URL}/api/repairs/{rid}", timeout=60)

    def test_tax_21(self, owner_client, card_id):
        r = owner_client.put(f"{BASE_URL}/api/repairs/{card_id}",
                             json={"labor_charge": 100.0, "tax_rate": 21}, timeout=60)
        assert r.status_code == 200, r.text[:300]
        c = r.json()
        assert c["grand_total"] == 100.0
        assert c["tax_rate"] == 21
        assert c["tax_amount"] == 21.0
        assert c["total_with_tax"] == 121.0
        g = owner_client.get(f"{BASE_URL}/api/repairs/{card_id}", timeout=60).json()
        assert g["tax_amount"] == 21.0 and g["total_with_tax"] == 121.0

    def test_tax_9(self, owner_client, card_id):
        r = owner_client.put(f"{BASE_URL}/api/repairs/{card_id}",
                             json={"labor_charge": 200.0, "tax_rate": 9}, timeout=60)
        assert r.status_code == 200, r.text[:300]
        c = r.json()
        assert c["grand_total"] == 200.0
        assert c["tax_amount"] == 18.0
        assert c["total_with_tax"] == 218.0

    def test_tax_zero(self, owner_client, card_id):
        r = owner_client.put(f"{BASE_URL}/api/repairs/{card_id}",
                             json={"tax_rate": 0}, timeout=60)
        assert r.status_code == 200, r.text[:300]
        c = r.json()
        assert c["tax_amount"] == 0.0
        assert c["total_with_tax"] == c["grand_total"]


# ================= Cash movements =================
class TestCashMovements:
    def test_create_mirror_list_delete(self, owner_client, method_id):
        payload = {"date": TODAY, "direction": "OUT", "amount": 37.55,
                   "payment_method_id": method_id, "category": "expense", "note": "TEST_movement"}
        r = owner_client.post(f"{BASE_URL}/api/cash-movements", json=payload, timeout=60)
        assert r.status_code == 200, r.text[:400]
        mv = r.json()
        assert mv["direction"] == "OUT" and mv["amount"] == 37.55
        assert mv["payment_method_id"] == method_id and mv["payment_method_name"]
        assert mv["category"] == "expense" and mv["note"] == "TEST_movement"
        mid = mv["id"]

        lr = owner_client.get(f"{BASE_URL}/api/cash-movements?date={TODAY}", timeout=60)
        assert lr.status_code == 200
        assert any(x["id"] == mid for x in lr.json())

        # mirrored ledger row on the payment method
        er = owner_client.get(f"{BASE_URL}/api/payment-entries?method_id={method_id}", timeout=60)
        assert er.status_code == 200, er.text[:300]
        rows = er.json()
        mirrored = [e for e in rows if e.get("source_id") == mid]
        assert mirrored, "No mirrored payment_entries row with source='cash_movement'"
        assert mirrored[0]["source"] == "cash_movement"
        assert mirrored[0]["amount"] == 37.55
        assert mirrored[0]["direction"] == "OUT"

        dr = owner_client.delete(f"{BASE_URL}/api/cash-movements/{mid}", timeout=60)
        assert dr.status_code == 200, dr.text[:300]
        lr2 = owner_client.get(f"{BASE_URL}/api/cash-movements?date={TODAY}", timeout=60)
        assert not any(x["id"] == mid for x in lr2.json())
        er2 = owner_client.get(f"{BASE_URL}/api/payment-entries?method_id={method_id}", timeout=60)
        assert er2.status_code == 200
        assert not [e for e in er2.json() if e.get("source_id") == mid], "ledger row not deleted"

    def test_bad_direction_rejected(self, owner_client):
        r = owner_client.post(f"{BASE_URL}/api/cash-movements",
                              json={"direction": "SIDEWAYS", "amount": 5}, timeout=60)
        assert r.status_code in (400, 422), f"{r.status_code} {r.text[:200]}"

    def test_negative_amount_rejected(self, owner_client):
        r = owner_client.post(f"{BASE_URL}/api/cash-movements",
                              json={"direction": "IN", "amount": -10}, timeout=60)
        assert r.status_code == 422, f"{r.status_code} {r.text[:200]}"

    def test_unknown_method_404(self, owner_client):
        r = owner_client.post(f"{BASE_URL}/api/cash-movements",
                              json={"direction": "IN", "amount": 10,
                                    "payment_method_id": "nope-nope"}, timeout=60)
        assert r.status_code == 404, f"{r.status_code} {r.text[:200]}"

    def test_delete_missing_404(self, owner_client):
        assert owner_client.delete(f"{BASE_URL}/api/cash-movements/nope", timeout=60).status_code == 404

    def test_requires_auth(self):
        assert requests.get(f"{BASE_URL}/api/cash-movements", timeout=60).status_code == 401
        assert requests.post(f"{BASE_URL}/api/cash-movements",
                             json={"direction": "IN", "amount": 1}, timeout=60).status_code == 401

    def test_staff_can_read_and_create(self, staff_client, owner_client):
        r = staff_client.post(f"{BASE_URL}/api/cash-movements",
                              json={"direction": "IN", "amount": 12.5, "category": "deposit",
                                    "note": "TEST_staff_mv"}, timeout=60)
        assert r.status_code == 200, r.text[:300]
        mid = r.json()["id"]
        assert staff_client.get(f"{BASE_URL}/api/cash-movements", timeout=60).status_code == 200
        # NOTE: delete is not owner-gated -> record actual behaviour
        d = staff_client.delete(f"{BASE_URL}/api/cash-movements/{mid}", timeout=60)
        owner_client.delete(f"{BASE_URL}/api/cash-movements/{mid}", timeout=60)
        assert d.status_code in (200, 403), d.status_code


# ================= Cash register aggregation =================
class TestCashRegisterManual:
    def test_manual_totals_and_netflow(self, owner_client, method_id):
        before = owner_client.get(f"{BASE_URL}/api/cash-register?date={TODAY}", timeout=60)
        assert before.status_code == 200, before.text[:300]
        b = before.json()
        for k in ("manual_in", "manual_out", "manual_movements", "net_flow"):
            assert k in b, f"missing {k}"
        r = owner_client.post(f"{BASE_URL}/api/cash-movements",
                              json={"date": TODAY, "direction": "IN", "amount": 50.0,
                                    "payment_method_id": method_id, "category": "deposit",
                                    "note": "TEST_till_in"}, timeout=60)
        assert r.status_code == 200, r.text[:300]
        mid = r.json()["id"]
        try:
            after = owner_client.get(f"{BASE_URL}/api/cash-register?date={TODAY}", timeout=60).json()
            assert round(after["manual_in"] - b["manual_in"], 2) == 50.0
            assert round(after["net_flow"] - b["net_flow"], 2) == 50.0
            assert any(m["id"] == mid for m in after["manual_movements"])
        finally:
            owner_client.delete(f"{BASE_URL}/api/cash-movements/{mid}", timeout=60)

    def test_manual_out_reduces_netflow(self, owner_client):
        b = owner_client.get(f"{BASE_URL}/api/cash-register?date={TODAY}", timeout=60).json()
        r = owner_client.post(f"{BASE_URL}/api/cash-movements",
                              json={"date": TODAY, "direction": "OUT", "amount": 20.0,
                                    "category": "withdrawal", "note": "TEST_till_out"}, timeout=60)
        assert r.status_code == 200, r.text[:300]
        mid = r.json()["id"]
        try:
            a = owner_client.get(f"{BASE_URL}/api/cash-register?date={TODAY}", timeout=60).json()
            assert round(a["manual_out"] - b["manual_out"], 2) == 20.0
            assert round(a["net_flow"] - b["net_flow"], 2) == -20.0
        finally:
            owner_client.delete(f"{BASE_URL}/api/cash-movements/{mid}", timeout=60)


# ================= Excel exports =================
def _sheet_rows(content, sheet=None):
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content))
    ws = wb[sheet] if sheet else wb.active
    return wb, ws, [[c.value for c in row] for row in ws.iter_rows()]


class TestExcelExports:
    def test_inventory_excel(self, owner_client):
        r = owner_client.get(f"{BASE_URL}/api/reports/inventory/excel", timeout=120)
        assert r.status_code == 200, r.text[:300]
        assert r.headers["content-type"] == XLSX_CT
        assert zipfile.is_zipfile(io.BytesIO(r.content))
        wb, ws, rows = _sheet_rows(r.content)
        assert ws.title == "Inventory"
        assert rows[0] == ["SKU", "Barcode", "Name", "Category", "Location", "Qty",
                           "Reorder", "Cost €", "Selling €", "Stock value €"]
        inv = owner_client.get(f"{BASE_URL}/api/inventory", timeout=90).json()
        items = inv if isinstance(inv, list) else inv.get("items", [])
        assert len(rows) - 1 == len(items), f"rows={len(rows)-1} items={len(items)}"

    def test_invoices_excel(self, owner_client):
        r = owner_client.get(f"{BASE_URL}/api/reports/invoices/excel", timeout=120)
        assert r.status_code == 200, r.text[:300]
        assert r.headers["content-type"] == XLSX_CT
        wb, ws, rows = _sheet_rows(r.content)
        hdr = rows[0]
        for col in ("Invoice #", "Subtotal €", "Tax €", "Total €", "Status"):
            assert col in hdr, f"missing column {col}"

    def test_profit_excel_two_sheets(self, owner_client):
        r = owner_client.get(f"{BASE_URL}/api/reports/profit/excel?start=2025-01-01&end={TODAY}",
                             timeout=120)
        assert r.status_code == 200, r.text[:300]
        assert r.headers["content-type"] == XLSX_CT
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(r.content))
        assert wb.sheetnames == ["Summary", "By part"], wb.sheetnames
        summary = [[c.value for c in row] for row in wb["Summary"].iter_rows()]
        metrics = {row[0] for row in summary}
        for m in ("From", "To", "Total revenue €", "Gross profit €", "Margin %"):
            assert m in metrics, f"missing metric {m}"

    def test_cash_register_excel(self, owner_client):
        r = owner_client.get(f"{BASE_URL}/api/reports/cash-register/excel?date={TODAY}", timeout=120)
        assert r.status_code == 200, r.text[:300]
        assert r.headers["content-type"] == XLSX_CT
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(r.content))
        assert f"Till {TODAY}" in wb.sheetnames and "Summary" in wb.sheetnames, wb.sheetnames
        summ = {row[0].value: row[1].value for row in wb["Summary"].iter_rows(min_row=2)}
        for k in ("Date", "Paid invoices", "Revenue €", "Manual IN €", "Manual OUT €", "Net €"):
            assert k in summ, f"missing {k}"
        assert summ["Date"] == TODAY

    def test_excel_requires_auth(self):
        for p in ("inventory", "invoices", "profit"):
            assert requests.get(f"{BASE_URL}/api/reports/{p}/excel",
                                timeout=60).status_code == 401, p
        assert requests.get(f"{BASE_URL}/api/reports/cash-register/excel",
                            timeout=60).status_code == 401


# ================= Settings default_tax_rate =================
class TestSettingsTaxRate:
    def test_persist_tax_rate(self, owner_client):
        cur = owner_client.get(f"{BASE_URL}/api/settings", timeout=60)
        assert cur.status_code == 200, cur.text[:300]
        original = cur.json().get("default_tax_rate", 21)
        assert "default_tax_rate" in cur.json()
        try:
            for val in (9, 21):
                body = dict(cur.json())
                body["default_tax_rate"] = val
                r = owner_client.put(f"{BASE_URL}/api/settings", json=body, timeout=60)
                assert r.status_code == 200, r.text[:300]
                assert r.json()["default_tax_rate"] == val
                g = owner_client.get(f"{BASE_URL}/api/settings", timeout=60).json()
                assert g["default_tax_rate"] == val
        finally:
            body = dict(cur.json())
            body["default_tax_rate"] = original
            owner_client.put(f"{BASE_URL}/api/settings", json=body, timeout=60)

    def test_staff_cannot_update_settings(self, staff_client):
        r = staff_client.put(f"{BASE_URL}/api/settings", json={"default_tax_rate": 5}, timeout=60)
        assert r.status_code == 403, f"{r.status_code} {r.text[:200]}"


# ================= Regression =================
class TestRegression:
    def test_core_endpoints_ok(self, owner_client):
        for path in ("/api/invoices", "/api/repairs", f"/api/cash-register?date={TODAY}",
                     "/api/inventory", "/api/customers", "/api/payment-methods",
                     "/api/settings", "/api/reports/profit"):
            r = owner_client.get(f"{BASE_URL}{path}", timeout=90)
            assert r.status_code == 200, f"{path} -> {r.status_code} {r.text[:200]}"

    def test_backup_collections_include_cash_movements(self):
        src = open("/app/backend/backup.py").read()
        assert '"cash_movements"' in src, "cash_movements not in MANAGED_COLLECTIONS"

    def test_no_mongo_id_leak(self, owner_client):
        for path in ("/api/repairs", "/api/cash-movements", "/api/invoices"):
            r = owner_client.get(f"{BASE_URL}{path}", timeout=90)
            assert r.status_code == 200
            assert '"_id"' not in r.text, f"_id leaked in {path}"
